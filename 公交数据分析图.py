#Date
from numpy import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook


data=load_workbook('C:\\Users\\Administrator\\Desktop\\1.xlsx')
table=data.get_sheet_by_name('基础数据表')
for i in range(2,1560):
    Su=table.cell(row=i,column=17).value
    Fu=table.cell(row=i,column=24).value
    Jin=table.cell(row=i,column=20).value
    Wei=table.cell(row=i,column=21).value
    xin=table.cell(row=i,column=12).value
    if Su == '已安装':
        if Fu == '正常' or ' ':
            plt.scatter(Jin,Wei,alpha=0.3,color='b')
        else:
            plt.scatter(Jin,Wei,alpha=0.5,color='r')
    else:
        if xin == '是':
           plt.scatter(Jin,Wei,alpha=0.7,color='y')
        else:
           plt.scatter(Jin,Wei,alpha=1,color='k')
plt.xlim(121.3547,121.5530)
plt.ylim(31.1411,31.3252)
plt.xticks([])
plt.yticks([])
plt.show()
